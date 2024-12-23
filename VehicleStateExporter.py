import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
import os
import cv2

class VehicleStateExporter:
    def __init__(self, output_folder="output_excel", excel_file_name="vehicle_states.xlsx"):
        """
        Initialize the Excel exporter with output folder and file name.
        """
        self.output_folder = output_folder
        self.excel_file_path = os.path.join(self.output_folder, excel_file_name)
        os.makedirs(self.output_folder, exist_ok=True)  # Ensure output folder exists

        # Initialize the Excel workbook and worksheet
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "Vehicle States"

        # Set up headers
        headers = ["Tracker ID", "Max Speed (km/h)", "Number Plate", "Vehicle Image"]
        for col_num, header in enumerate(headers, 1):
            self.sheet.cell(row=1, column=col_num, value=header)

        self.row_num = 2  # Start populating data from row 2

    def add_vehicle_state(self, tracker_id, max_speed, number_plate, image):
        """
        Add vehicle state information and image to the Excel file.
        
        Args:
            tracker_id (int): The ID of the tracked vehicle.
            max_speed (float): The maximum speed of the vehicle.
            number_plate (str): The detected license plate text.
            image (np.ndarray): Cropped vehicle image in OpenCV (BGR format).
        """
        # Save the vehicle image temporarily
        image_path = os.path.join(self.output_folder, f"vehicle_{tracker_id}.jpg")
        if image is not None:
            cv2.imwrite(image_path, image)
        else:
            image_path = None

        # Add vehicle state information to Excel
        self.sheet.cell(row=self.row_num, column=1, value=tracker_id)  # Tracker ID
        self.sheet.cell(row=self.row_num, column=2, value=round(max_speed, 2))  # Max Speed
        self.sheet.cell(row=self.row_num, column=3, value=number_plate or "N/A")  # Number Plate

        # Insert the vehicle image
        if image_path:
            img = OpenpyxlImage(image_path)
            img.width, img.height = 80, 50  # Resize image to fit Excel cell
            self.sheet.add_image(img, f"D{self.row_num}")  # Add image to column D

        self.row_num += 1  # Move to the next row

    def save_excel(self):
        """
        Save the Excel workbook to the specified file path.
        """
        # Adjust column widths for better readability
        self.sheet.column_dimensions["A"].width = 15
        self.sheet.column_dimensions["B"].width = 20
        self.sheet.column_dimensions["C"].width = 25
        self.sheet.column_dimensions["D"].width = 20

        # Save the Excel file
        self.workbook.save(self.excel_file_path)
        print(f"Vehicle states saved to {self.excel_file_path}")
